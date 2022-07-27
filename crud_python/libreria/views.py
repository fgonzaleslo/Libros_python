from cgitb import html
from csv import excel
from datetime import datetime
#from crypt import methods
from distutils.command.upload import upload
#from crypt import methods
from fileinput import filename
from importlib.resources import path
from io import BytesIO
from msilib.schema import ListView
from multiprocessing import context
from pickle import OBJ
from poplib import POP3_SSL_PORT
from re import template
from unittest import result
from django.views.generic import ListView
from pipes import Template
from tokenize import Name
from urllib import response
import django
from django.shortcuts import render, redirect
from django.http import HttpResponse
from faker import Faker
from numpy import identity
import openpyxl
from requests import request
from .models import Libro,Autor
from .forms import AutorForm, LibroForm
#Excel
from openpyxl import Workbook
from django.views.generic import TemplateView
#PDF
from django.template.loader import get_template
from xhtml2pdf import pisa
from io import BytesIO
#Busqueda
from django.db.models import Q
#paginas
from django.core.paginator import Paginator
#importar
import os
import pandas as pd
from django.conf import settings
import os
from django.core.files.storage import FileSystemStorage
from django.conf import settings

import xlwt
import datetime

import csv


# Create your views here.
#request para hacer la solicitud de la vista, que al final responde con un texto
def inicio(request):
    return render(request, 'paginas/inicio.html')

def nosotros(request):
    return render(request, 'paginas/nosotros.html')

def libros(request):
    libros = Libro.objects.all()
    paginator = Paginator(libros,5)
    pagina = request.GET.get("page") or 1
    libros = paginator.get_page(pagina)
    pagina_actual = int(pagina)
    paginas = range(1, libros.paginator.num_pages + 1)

    #print(libros)
    queryset = request.GET.get("buscar")
    #libros = Libro.objects.filter(titulo = True)
    if queryset:
        libros = Libro.objects.filter(
            Q(titulo__icontains = queryset)
        ).distinct()
    return render(request, 'libros/index.html',{'libros': libros, 'paginas': paginas, 'pagina_actual': pagina_actual})

def crear(request):
    #identifiamos todos los elementos que nos envian desde el formulario
    formulario = LibroForm(request.POST or None)
    if formulario.is_valid():
        formulario.save()
        return redirect('libros')
    return render(request, 'libros/crear.html',{'formulario': formulario})

def editar(request, id):
    libro = Libro.objects.get(id=id)
    formulario = LibroForm(request.POST or None, instance=libro)
    if formulario.is_valid() and request.POST:
        formulario.save()
        return redirect('libros')
    return render(request, 'libros/editar.html', {'formulario': formulario})

#función eliminar
def eliminar(request, id):
    libro = Libro.objects.get(id=id)
    libro.delete()
    return redirect('libros')


#funcion exportar PDF
def html_to_pdf(template_src, context_dict={}):
     template = get_template(template_src)
     html  = template.render(context_dict)
     result = BytesIO()
     pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
     if not pdf.err:
         return HttpResponse(result.getvalue(), content_type='application/pdf')
     return None

class GeneratePdf(TemplateView):
    def get(self, request, *args, **kwargs):
        libros = Libro.objects.all()
        data = {
             'libros' : libros
         }
         
        # getting the template
        pdf = html_to_pdf('libros/reportepdf.html', data)
         
         # rendering the template
        return HttpResponse(pdf, content_type='application/pdf')

#Generar reporte Excel
class ReporteExel(TemplateView):
    def get(self,request,*args,**kwargs):
        libro = Libro.objects.all()
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE DJANGO'

        ws.merge_cells('B1:E1')

        ws['B3'] = 'ID'
        ws['C3'] = 'TITULO'
        ws['D3'] = 'DESCRIPCION'

        cont = 3

        for librosv in libro:
            ws.cell(row = cont, column=2).value = librosv.id
            ws.cell(row = cont, column=3).value = librosv.titulo
            ws.cell(row = cont, column=4).value = librosv.descripcion
            cont+=1

        nombre_archivo = "ReporteLibro.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response

def exportar_excel_reporte(request):
    response=HttpResponse(content_type='aplication/ms-excel')
    response['Content-Disposition']= 'attachment; filename=Reporte de libros' + \
        str(datetime.datetime.now())+'.xls'

    wb = xlwt.Workbook(encoding='utf-8')
    ws=wb.add_sheet('Reporte de libros')
    libroxl_nun = 0
    font_style=xlwt.XFStyle()
    font_style.font.bold = True

    columns = ['ID','TITULO','DESCRIPCION']

    for col_num in range(len(columns)):
        ws.write(libroxl_nun, col_num, columns[col_num], font_style)

    font_style = xlwt.XFStyle()

    libros = Libro.objects.all().values_list('id', 'titulo', 'descripcion')

    for libroxl in libros:
        libroxl_nun +=1

        for col_num in range(len(libroxl)):
            ws.write(libroxl_nun, col_num, str(libroxl[col_num]), font_style)
    wb.save(response)

    return response


    


def importar_libro(request):
    context={}

    if request.method == 'POST':
        uploaded_file = request.FILES['document']

        print(uploaded_file)
    
        if uploaded_file.name.endswith('.csv'):
            savefile = FileSystemStorage()

            name = savefile.save(uploaded_file.name, uploaded_file)

            d = os.getcwd()
            file_diretory = d+'\media\\'+name

    return render(request, 'libros/importar.html')

#Autores
def autores(request):
    autores = Autor.objects.all()
    #print(libros)
    return render(request, 'autores/autores.html',{'autores': autores})

def crear_autor(request):
    #identifiamos todos los elementos que nos envian desde el formulario
    formulario = AutorForm(request.POST or None)
    if formulario.is_valid():
        formulario.save()
        return redirect('autores')
    return render(request, 'autores/crear.html',{'formulario': formulario})

#función eliminar
def eliminar_autor(request, id):
    autores = Autor.objects.get(id=id)
    autores.delete()
    return redirect('autores')


def editar_autor(request, id):
    autores = Autor.objects.get(id=id)
    formulario = AutorForm(request.POST or None, instance=autores)
    if formulario.is_valid() and request.POST:
        formulario.save()
        return redirect('autores')
    return render(request, 'autores/editar.html', {'formulario': formulario})

def exportar_csv_reporte(request):
    response=HttpResponse(content_type='text/csv')
    response['Content-Disposition']= 'attachment; filename=Reporte de libros' + \
        str(datetime.datetime.now())+'.csv'

    writer = csv.writer(response)
    writer.writerow(['ID','NOMBRES','APELLIDOS','DNI','DIRECCION','LIBRO'])

    autores = Autor.objects.all()

    for autor in autores:
        writer.writerow([autor.id,autor.nombres,autor.apellidos,autor.dni,autor.direccion,autor.libro_id])

    return response




    